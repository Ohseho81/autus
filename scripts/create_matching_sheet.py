#!/usr/bin/env python3
"""온리쌤 개발 현황 매칭표 생성"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Sheet 1: 기능 매칭표
ws1 = wb.active
ws1.title = "기능 매칭표"

# 헤더 스타일
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

# 헤더
headers1 = ["기능명", "우선순위", "개발상태", "완성도(%)", "필요작업", "기술스택", "비고"]
ws1.append(headers1)

for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(1, col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 데이터
data1 = [
    # P0 기능
    ["미수금 자동 추적", "P0", "부분완료", "60", "대시보드 UI, 실시간 조회", "Supabase, Next.js", "결제 테이블 활용"],
    ["학부모 자동 알림", "P0", "완료", "90", "테스트 및 배포", "카카오톡 API, Python", "kakao_notification.py"],
    ["수업 결과 로그", "P0", "미착수", "0", "테이블 생성, 입력폼, 알림 연동", "Supabase, React", "class_logs 테이블 필요"],

    # P1 기능
    ["수업 스케줄 관리", "P1", "미착수", "0", "테이블 생성, 충돌 감지 로직", "Supabase, Algorithm", "schedules 테이블 필요"],
    ["데이터 자동 동기화", "P1", "완료", "95", "자동 스케줄 설정", "Python, Plugin", "플러그인 완성됨"],

    # P2 기능
    ["출석 관리", "P2", "미착수", "0", "QR 체크인, 출결 기록", "QR Library, DB", ""],
    ["매출 리포트", "P2", "미착수", "0", "데이터 집계, 차트 생성", "Pandas, Recharts", ""],

    # 인프라
    ["Supabase DB", "인프라", "부분완료", "60", "신규 테이블 4개 생성", "PostgreSQL", "students, memberships, payments"],
    ["카카오톡 연동", "인프라", "완료", "100", "프로덕션 배포", "Kakao API", "토큰 설정 완료"],
    ["플러그인 시스템", "인프라", "완료", "100", "설치 및 배포", "Claude Plugin", "autus-supabase-uploader.plugin"],
]

for row in data1:
    ws1.append(row)

# 조건부 서식 (완성도에 따른 색상)
status_colors = {
    "완료": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "부분완료": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "미착수": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

for row in range(2, ws1.max_row + 1):
    status = ws1.cell(row, 3).value
    if status in status_colors:
        ws1.cell(row, 3).fill = status_colors[status]

# 컬럼 너비 조정
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 35
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 30

# Sheet 2: 데이터베이스 현황
ws2 = wb.create_sheet("데이터베이스 현황")

headers2 = ["테이블명", "상태", "컬럼수", "예상 데이터수", "용도", "비고"]
ws2.append(headers2)

for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(1, col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

data2 = [
    # 기존 테이블
    ["students", "구축완료", "8", "781", "학생 기본 정보", "업로드 준비 완료"],
    ["memberships", "구축완료", "10", "853", "회원권/수업 정보", "업로드 준비 완료"],
    ["payments", "구축완료", "7", "776", "결제/미수금", "업로드 준비 완료"],

    # 신규 테이블
    ["class_logs", "설계완료", "8", "예상 10,000+/년", "수업 결과 기록", "P0 - 즉시 생성 필요"],
    ["schedules", "설계완료", "8", "예상 50-100", "정규 수업 시간표", "P1 - 2주내 생성"],
    ["classes", "설계완료", "6", "예상 5,000+/년", "실제 수업 세션", "P1 - 2주내 생성"],
    ["notifications", "설계완료", "7", "예상 30,000+/년", "알림 발송 이력", "P1 - 추적용"],
]

for row in data2:
    ws2.append(row)

# 상태별 색상
db_status_colors = {
    "구축완료": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "설계완료": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

for row in range(2, ws2.max_row + 1):
    status = ws2.cell(row, 2).value
    if status in db_status_colors:
        ws2.cell(row, 2).fill = db_status_colors[status]

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 25
ws2.column_dimensions['F'].width = 30

# Sheet 3: 구현 로드맵
ws3 = wb.create_sheet("구현 로드맵")

headers3 = ["주차", "기간", "목표", "세부 작업", "상태", "완료도(%)"]
ws3.append(headers3)

for col_num, header in enumerate(headers3, 1):
    cell = ws3.cell(1, col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

data3 = [
    # Week 1-2: P0 완성
    ["Week 1-2", "2026-02-14 ~ 02-28", "P0 기능 완성",
     "1. class_logs 테이블 생성\n2. 수업 결과 입력 폼 (모바일)\n3. 학부모 자동 알림 연동\n4. 미수금 대시보드",
     "진행중", "30"],

    # Week 3-4: P1 완성
    ["Week 3-4", "2026-03-01 ~ 03-14", "P1 기능 완성",
     "1. schedules/classes 테이블 생성\n2. 수업 자동 생성 로직\n3. 스케줄 UI (캘린더)\n4. 플러그인 자동화 스케줄",
     "미착수", "0"],

    # Week 5-6: 테스트 & 안정화
    ["Week 5-6", "2026-03-15 ~ 03-28", "테스트 & 안정화",
     "1. 전체 시스템 통합 테스트\n2. 성능 최적화\n3. 버그 수정\n4. 사용자 교육",
     "미착수", "0"],

    # Week 7-8: P2 개선
    ["Week 7-8", "2026-03-29 ~ 04-11", "P2 기능 추가",
     "1. QR 출석 체크인\n2. 매출 리포트\n3. 예측 분석",
     "미착수", "0"],
]

for row_data in data3:
    ws3.append(row_data)

# 로드맵 상태별 색상
roadmap_colors = {
    "완료": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "진행중": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "미착수": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

for row in range(2, ws3.max_row + 1):
    status = ws3.cell(row, 5).value
    if status in roadmap_colors:
        ws3.cell(row, 5).fill = roadmap_colors[status]
    # 텍스트 wrap
    ws3.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 50
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12

# Sheet 4: 파일 현황
ws4 = wb.create_sheet("파일 현황")

headers4 = ["파일명", "유형", "상태", "용도", "위치"]
ws4.append(headers4)

for col_num, header in enumerate(headers4, 1):
    cell = ws4.cell(1, col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

data4 = [
    ["students.csv", "데이터", "준비완료", "781명 학생 데이터", "autus/"],
    ["upload_students.py", "스크립트", "완료", "Supabase 업로드", "autus/"],
    ["kakao_notification.py", "스크립트", "완료", "카카오톡 자동 알림", "autus/"],
    ["autus-supabase-uploader.plugin", "플러그인", "완료", "데이터 자동 업로드", "autus/"],
    ["ONLYSSAM_SPEC.md", "문서", "완료", "시스템 스펙 정의", "autus/"],
    ["README_KAKAO.md", "문서", "완료", "카카오톡 연동 가이드", "autus/"],
    ["supabase_upload_colab.ipynb", "노트북", "완료", "Google Colab 업로드", "autus/"],
]

for row in data4:
    ws4.append(row)

file_colors = {
    "완료": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "준비완료": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

for row in range(2, ws4.max_row + 1):
    status = ws4.cell(row, 3).value
    if status in file_colors:
        ws4.cell(row, 3).fill = file_colors[status]

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 30
ws4.column_dimensions['E'].width = 20

# 저장
wb.save('/sessions/modest-bold-einstein/mnt/autus/온리쌤_개발현황_매칭표.xlsx')
print("✅ 온리쌤_개발현황_매칭표.xlsx 생성 완료!")
