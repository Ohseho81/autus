#!/usr/bin/env python3
"""결제선생 통합 설계 Excel 생성"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# 색상 정의
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HIGHLIGHT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD_FONT = Font(bold=True)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(sheet, widths):
    for idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

def add_header(sheet, row, values, fill=HEADER_FILL, font=HEADER_FONT):
    for col, value in enumerate(values, 1):
        cell = sheet.cell(row, col, value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

# ==================== Sheet 1: Overview ====================
ws1 = wb.create_sheet('Overview')
set_column_widths(ws1, [25, 50, 15])

ws1['A1'] = '💳 결제선생 → AUTUS 통합 설계'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws1.merge_cells('A1:C1')

ws1['A3'] = '프로젝트 정보'
ws1['A3'].font = BOLD_FONT
ws1['A3'].fill = SUBHEADER_FILL

ws1['A4'] = '목표'
ws1['B4'] = '결제선생 청구·수납 시스템을 온리쌤 Supabase에 통합'
ws1['A5'] = '범위'
ws1['B5'] = '8개 핵심 엔티티 동기화 + 이벤트 로깅'
ws1['A6'] = '기간'
ws1['B6'] = 'Week 2-3 (2주)'
ws1['A7'] = '예상 비용'
ws1['B7'] = '월 40만원 (매출의 4%)'

ws1['A9'] = '현재 상태'
ws1['A9'].font = BOLD_FONT
ws1['A9'].fill = SUBHEADER_FILL

data = [
    ['항목', '상태', '진행률'],
    ['기존 Supabase 스키마', '✅ 완료', '100%'],
    ['신규 테이블 설계', '✅ 완료', '100%'],
    ['신규 테이블 생성', '⏳ 대기', '0%'],
    ['API 엔드포인트 개발', '⏳ 대기', '0%'],
    ['결제선생 연동', '⏳ 대기', '0%'],
    ['카카오톡 알림 추가', '⏳ 대기', '0%'],
]
add_header(ws1, 10, data[0])
for i, row in enumerate(data[1:], 11):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(i, j, val)
        cell.border = BORDER
        if '100%' in val:
            cell.fill = SUCCESS_FILL
        elif '0%' in val:
            cell.fill = WARNING_FILL

ws1['A18'] = '핵심 변경사항'
ws1['A18'].font = BOLD_FONT
ws1['A18'].fill = SUBHEADER_FILL

ws1['A19'] = '신규 테이블 4개'
ws1['B19'] = 'invoices, payment_transactions, cash_receipts, business_settings'
ws1['A20'] = '기존 테이블 확장'
ws1['B20'] = 'payments에 invoice_id, latest_transaction_id 추가'
ws1['A21'] = '신규 VIEW 2개'
ws1['B21'] = 'daily_sales_report, invoice_status_summary'
ws1['A22'] = '신규 API 7개'
ws1['B22'] = '청구서 생성/발송, 결제 처리, 현금영수증, 매출 보고서 등'

# ==================== Sheet 2: Schema Mapping ====================
ws2 = wb.create_sheet('Schema Mapping')
set_column_widths(ws2, [20, 25, 25, 30, 15])

ws2['A1'] = '결제선생 엔티티 → Supabase 매핑'
ws2['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws2.merge_cells('A1:E1')

data = [
    ['결제선생 엔티티', 'Supabase 테이블', '매핑 방식', '주요 컬럼', '상태'],
    ['학생 데이터', 'profiles', '기존 활용', 'name, phone, type', '✅ 완료'],
    ['청구서 데이터', 'invoices', '신규 생성', 'invoice_number, items, status', '⏳ 설계'],
    ['결제 내역', 'payment_transactions', '신규 생성', 'amount, card_company, approval_number', '⏳ 설계'],
    ['발송·수납 내역', 'invoices', '신규 생성', 'sent_at, paid_at, status', '⏳ 설계'],
    ['현금영수증', 'cash_receipts', '신규 생성', 'approval_number, recipient_number', '⏳ 설계'],
    ['매출 보고서', 'daily_sales_report (VIEW)', 'VIEW 생성', 'sale_date, total_sales, card_sales', '⏳ 설계'],
    ['출결 데이터', 'bookings + attendance', '기존 활용', 'booking_date, status', '✅ 완료'],
    ['사업장 정보', 'business_settings', '신규 생성', 'business_name, pg_provider', '⏳ 설계'],
]

add_header(ws2, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i, j, val)
        cell.border = BORDER
        if '✅' in val:
            cell.fill = SUCCESS_FILL
        elif '⏳' in val:
            cell.fill = WARNING_FILL

# ==================== Sheet 3: New Tables ====================
ws3 = wb.create_sheet('New Tables')
set_column_widths(ws3, [25, 20, 15, 40])

ws3['A1'] = '신규 테이블 상세'
ws3['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws3.merge_cells('A1:D1')

# invoices
ws3['A3'] = '1. invoices (청구서)'
ws3['A3'].font = BOLD_FONT
ws3['A3'].fill = SUBHEADER_FILL
ws3.merge_cells('A3:D3')

data = [
    ['컬럼명', '타입', '필수', '설명'],
    ['invoice_number', 'TEXT', 'Y', '청구서 번호 (INV-20260214-001)'],
    ['student_id', 'UUID', 'Y', '학생 ID (profiles 참조)'],
    ['items', 'JSONB', 'Y', '청구 항목 [{name, amount, qty}]'],
    ['total_amount', 'INTEGER', 'Y', '총 청구 금액'],
    ['final_amount', 'INTEGER', 'Y', '최종 금액 (할인 적용 후)'],
    ['status', 'TEXT', 'Y', 'draft/sent/paid/partial/overdue/cancelled'],
    ['sent_at', 'TIMESTAMPTZ', 'N', '발송 시각'],
    ['paid_amount', 'INTEGER', 'N', '수납 금액'],
    ['due_date', 'DATE', 'N', '납부 기한'],
]
add_header(ws3, 4, data[0])
for i, row in enumerate(data[1:], 5):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(i, j, val)
        cell.border = BORDER

# payment_transactions
ws3['A16'] = '2. payment_transactions (결제 내역)'
ws3['A16'].font = BOLD_FONT
ws3['A16'].fill = SUBHEADER_FILL
ws3.merge_cells('A16:D16')

data = [
    ['컬럼명', '타입', '필수', '설명'],
    ['invoice_id', 'UUID', 'Y', '청구서 ID'],
    ['transaction_id', 'TEXT', 'Y', 'PG사 거래 고유번호'],
    ['approval_number', 'TEXT', 'N', '승인번호'],
    ['amount', 'INTEGER', 'Y', '결제 금액'],
    ['fee', 'INTEGER', 'N', '수수료'],
    ['payment_method', 'TEXT', 'Y', 'card/cash/transfer/virtual_account'],
    ['card_company', 'TEXT', 'N', '매입사 (신한/국민/삼성 등)'],
    ['status', 'TEXT', 'Y', 'pending/completed/failed/cancelled'],
    ['paid_at', 'TIMESTAMPTZ', 'Y', '결제 시각'],
]
add_header(ws3, 17, data[0])
for i, row in enumerate(data[1:], 18):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(i, j, val)
        cell.border = BORDER

# cash_receipts
ws3['A28'] = '3. cash_receipts (현금영수증)'
ws3['A28'].font = BOLD_FONT
ws3['A28'].fill = SUBHEADER_FILL
ws3.merge_cells('A28:D28')

data = [
    ['컬럼명', '타입', '필수', '설명'],
    ['transaction_id', 'UUID', 'Y', '결제 내역 ID'],
    ['receipt_type', 'TEXT', 'Y', 'income(소득공제)/expenditure(지출증빙)'],
    ['recipient_number', 'TEXT', 'Y', '휴대폰 번호 or 사업자번호'],
    ['approval_number', 'TEXT', 'Y', '국세청 승인번호'],
    ['issued_at', 'TIMESTAMPTZ', 'Y', '발급 시각'],
    ['status', 'TEXT', 'Y', 'issued/cancelled'],
]
add_header(ws3, 29, data[0])
for i, row in enumerate(data[1:], 30):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(i, j, val)
        cell.border = BORDER

# business_settings
ws3['A37'] = '4. business_settings (사업장 정보)'
ws3['A37'].font = BOLD_FONT
ws3['A37'].fill = SUBHEADER_FILL
ws3.merge_cells('A37:D37')

data = [
    ['컬럼명', '타입', '필수', '설명'],
    ['business_name', 'TEXT', 'Y', '온리쌤배구아카데미'],
    ['business_number', 'TEXT', 'N', '사업자등록번호'],
    ['enabled_payment_methods', 'JSONB', 'N', '["card", "transfer"]'],
    ['pg_provider', 'TEXT', 'N', '결제선생/토스페이먼츠 등'],
    ['card_fee_rate', 'DECIMAL', 'N', '카드 수수료율 (기본 3.3%)'],
    ['auto_send_invoice', 'BOOLEAN', 'N', '자동 청구서 발송 여부'],
    ['auto_send_day', 'INTEGER', 'N', '매월 X일 발송 (기본 1일)'],
]
add_header(ws3, 38, data[0])
for i, row in enumerate(data[1:], 39):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(i, j, val)
        cell.border = BORDER

# ==================== Sheet 4: API Endpoints ====================
ws4 = wb.create_sheet('API Endpoints')
set_column_widths(ws4, [10, 35, 50, 15])

ws4['A1'] = '신규 FastAPI 엔드포인트 (7개)'
ws4['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws4.merge_cells('A1:D1')

data = [
    ['No', '엔드포인트', '기능', '상태'],
    ['1', 'POST /invoices', '청구서 생성 (학생, 항목, 금액, 납부기한)', '⏳'],
    ['2', 'POST /invoices/{id}/send', '청구서 발송 (카카오톡/SMS)', '⏳'],
    ['3', 'POST /payments/process', '결제 처리 (PG사 연동, 트랜잭션 기록)', '⏳'],
    ['4', 'POST /cash-receipts', '현금영수증 발급 (국세청 연동)', '⏳'],
    ['5', 'GET /reports/sales/daily', '일일 매출 보고서 (VIEW 조회)', '⏳'],
    ['6', 'GET /invoices/status', '월별 청구서 현황 (발송률, 수납률)', '⏳'],
    ['7', 'GET /invoices/unpaid', '미납 청구서 목록 (연체 포함)', '⏳'],
]
add_header(ws4, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(i, j, val)
        cell.border = BORDER
        if j == 4:
            cell.fill = WARNING_FILL

ws4['A12'] = '웹훅 엔드포인트'
ws4['A12'].font = BOLD_FONT
ws4['A12'].fill = SUBHEADER_FILL
ws4.merge_cells('A12:D12')

data = [
    ['No', '엔드포인트', '기능', '상태'],
    ['8', 'POST /webhooks/payment-teacher', '결제선생 웹훅 수신 (결제 완료 알림)', '⏳'],
]
add_header(ws4, 13, data[0])
for i, row in enumerate(data[1:], 14):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(i, j, val)
        cell.border = BORDER
        if j == 4:
            cell.fill = WARNING_FILL

# ==================== Sheet 5: Migration Plan ====================
ws5 = wb.create_sheet('Migration Plan')
set_column_widths(ws5, [10, 40, 15, 15, 15])

ws5['A1'] = '마이그레이션 계획 (Week 2-3)'
ws5['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws5.merge_cells('A1:E1')

ws5['A3'] = 'Phase 1: 스키마 확장 (Week 2, Day 1-2)'
ws5['A3'].font = BOLD_FONT
ws5['A3'].fill = SUBHEADER_FILL
ws5.merge_cells('A3:E3')

data = [
    ['No', '작업', '소요시간', '담당', '완료'],
    ['1', '신규 테이블 4개 생성 (invoices, payment_transactions 등)', '2시간', '개발자', '☐'],
    ['2', 'payments 테이블 컬럼 추가 (invoice_id, transaction_id)', '30분', '개발자', '☐'],
    ['3', 'VIEW 2개 생성 (daily_sales_report, invoice_status_summary)', '1시간', '개발자', '☐'],
    ['4', 'business_settings 초기 데이터 입력', '30분', '운영자', '☐'],
]
add_header(ws5, 4, data[0])
for i, row in enumerate(data[1:], 5):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

ws5['A10'] = 'Phase 2: 기존 데이터 마이그레이션 (Week 2, Day 3-4)'
ws5['A10'].font = BOLD_FONT
ws5['A10'].fill = SUBHEADER_FILL
ws5.merge_cells('A10:E10')

data = [
    ['No', '작업', '소요시간', '담당', '완료'],
    ['5', '기존 payments → invoices 변환 스크립트 작성', '3시간', '개발자', '☐'],
    ['6', '마이그레이션 스크립트 실행 (테스트 환경)', '1시간', '개발자', '☐'],
    ['7', '데이터 검증 (총액 일치, 레코드 수 확인)', '1시간', '운영자', '☐'],
    ['8', '프로덕션 마이그레이션', '2시간', '개발자', '☐'],
]
add_header(ws5, 11, data[0])
for i, row in enumerate(data[1:], 12):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

ws5['A17'] = 'Phase 3: 결제선생 API 연동 (Week 2, Day 5-7)'
ws5['A17'].font = BOLD_FONT
ws5['A17'].fill = SUBHEADER_FILL
ws5.merge_cells('A17:E17')

data = [
    ['No', '작업', '소요시간', '담당', '완료'],
    ['9', '결제선생 계정 생성 + API 키 발급', '30분', '운영자', '☐'],
    ['10', '결제선생 API 클라이언트 구현', '4시간', '개발자', '☐'],
    ['11', 'FastAPI 엔드포인트 7개 개발', '8시간', '개발자', '☐'],
    ['12', 'API 테스트 (청구서 발송, 결제 처리)', '2시간', '개발자', '☐'],
]
add_header(ws5, 18, data[0])
for i, row in enumerate(data[1:], 19):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

ws5['A24'] = 'Phase 4: 웹훅 + 카카오톡 (Week 3)'
ws5['A24'].font = BOLD_FONT
ws5['A24'].fill = SUBHEADER_FILL
ws5.merge_cells('A24:E24')

data = [
    ['No', '작업', '소요시간', '담당', '완료'],
    ['13', '웹훅 엔드포인트 개발', '3시간', '개발자', '☐'],
    ['14', '카카오톡 알림 템플릿 5개 추가', '2시간', '운영자', '☐'],
    ['15', '자동 청구서 발송 Edge Function', '4시간', '개발자', '☐'],
    ['16', '통합 테스트 (전체 플로우)', '4시간', '팀 전체', '☐'],
]
add_header(ws5, 25, data[0])
for i, row in enumerate(data[1:], 26):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

# ==================== Sheet 6: Cost Estimate ====================
ws6 = wb.create_sheet('Cost Estimate')
set_column_widths(ws6, [25, 20, 15, 30])

ws6['A1'] = '예상 비용 (월간)'
ws6['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws6.merge_cells('A1:D1')

data = [
    ['항목', '단가', '예상량', '월간 비용'],
    ['결제선생 이용료', '무료', '기본', 0],
    ['카드 결제 수수료', '0.8%', '1,000만원', 80000],
    ['현금영수증 발급', '20원/건', '1,000건', 20000],
    ['Supabase 스토리지', '무료', 'Free Tier', 0],
]
add_header(ws6, 3, data[0])
for i, row in enumerate(data[1:], 4):
    ws6.cell(i, 1, row[0]).border = BORDER
    ws6.cell(i, 2, row[1]).border = BORDER
    ws6.cell(i, 3, row[2]).border = BORDER
    cell = ws6.cell(i, 4, row[3])
    cell.border = BORDER
    cell.number_format = '#,##0'

ws6['A8'] = '합계'
ws6['A8'].font = BOLD_FONT
ws6['D8'] = '=SUM(D4:D7)'
ws6['D8'].font = BOLD_FONT
ws6['D8'].number_format = '#,##0'
ws6['D8'].fill = HIGHLIGHT_FILL

ws6['A10'] = '매출 대비 비용'
ws6['B10'] = '=D8/10000000'
ws6['B10'].number_format = '0.0%'

ws6['A12'] = '비용 상세'
ws6['A12'].font = BOLD_FONT
ws6['A12'].fill = SUBHEADER_FILL

ws6['A13'] = '- 카드 수수료는 매출에 비례하여 증가'
ws6['A14'] = '- 1,000만원 매출 기준 약 40만원 (4%)'
ws6['A15'] = '- 3,000만원 매출 시 약 120만원 (4%)'

# ==================== Sheet 7: Timeline ====================
ws7 = wb.create_sheet('Timeline')
set_column_widths(ws7, [15, 30, 20, 15])

ws7['A1'] = '개발 타임라인 (Week 2-3)'
ws7['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws7.merge_cells('A1:D1')

data = [
    ['주차', '작업', '담당', '상태'],
    ['Week 2 Day 1-2', '스키마 확장 (4개 테이블, 2개 VIEW)', '개발자', '⏳'],
    ['Week 2 Day 3-4', '기존 데이터 마이그레이션', '개발자', '⏳'],
    ['Week 2 Day 5-7', '결제선생 API 연동 + 7개 엔드포인트', '개발자', '⏳'],
    ['Week 3 Day 1-3', '웹훅 + 카카오톡 알림 5종', '개발자', '⏳'],
    ['Week 3 Day 4-5', '통합 테스트 + 버그 수정', '팀 전체', '⏳'],
]
add_header(ws7, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws7.cell(i, j, val)
        cell.border = BORDER
        if j == 4:
            cell.fill = WARNING_FILL

ws7['A10'] = '마일스톤'
ws7['A10'].font = BOLD_FONT
ws7['A10'].fill = SUBHEADER_FILL

data = [
    ['일정', '마일스톤', '산출물', ''],
    ['Week 2 종료', '기본 인프라 완성', 'Supabase 스키마 + API', ''],
    ['Week 3 종료', '통합 완료', '결제선생 연동 + 자동화', ''],
]
add_header(ws7, 11, data[0])
for i, row in enumerate(data[1:], 12):
    for j, val in enumerate(row, 1):
        cell = ws7.cell(i, j, val)
        cell.border = BORDER

# ==================== Sheet 8: Kakao Templates ====================
ws8 = wb.create_sheet('Kakao Templates')
set_column_widths(ws8, [25, 50, 20])

ws8['A1'] = '카카오톡 알림 템플릿 (5종 추가)'
ws8['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws8.merge_cells('A1:C1')

data = [
    ['템플릿명', '내용', '트리거'],
    ['청구서 발송', '2월 수업료 청구서가 발송되었습니다.\n금액: 200,000원\n납부기한: 2/28', 'invoices.sent_at'],
    ['결제 완료', '결제가 완료되었습니다.\n금액: 200,000원\n승인번호: 12345678', 'payment_transactions.paid_at'],
    ['미납 알림', '납부기한이 3일 남았습니다.\n미납금액: 200,000원', 'cron (매일)'],
    ['연체 알림', '납부기한이 7일 경과했습니다.\n미납금액: 200,000원', 'cron (매일)'],
    ['현금영수증 발급', '현금영수증이 발급되었습니다.\n승인번호: CR-20260214-001', 'cash_receipts.issued_at'],
]
add_header(ws8, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws8.cell(i, j, val)
        cell.border = BORDER
        if '\n' in str(val):
            cell.alignment = Alignment(wrap_text=True, vertical='top')

ws8['A10'] = '템플릿 승인 프로세스'
ws8['A10'].font = BOLD_FONT
ws8['A10'].fill = SUBHEADER_FILL

ws8['A11'] = '1. 카카오 비즈니스 채널 → 메시지 템플릿 관리'
ws8['A12'] = '2. 템플릿 5개 등록 (위 내용 참고)'
ws8['A13'] = '3. 승인 요청'
ws8['A14'] = '4. 승인 대기 (1-2 영업일)'
ws8['A15'] = '5. 승인 후 Solapi에서 template_id 확인'

wb.save('/sessions/modest-bold-einstein/mnt/autus/결제선생_통합_설계.xlsx')
print("✅ Excel 파일 생성 완료")
