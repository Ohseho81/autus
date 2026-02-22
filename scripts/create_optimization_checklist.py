#!/usr/bin/env python3
"""Supabase 최적화 체크리스트 Excel 생성"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
PHASE1_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PHASE2_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
PHASE3_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD_FONT = Font(bold=True)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(sheet, widths):
    for idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

def add_header(sheet, row, values, fill=HEADER_FILL, font=HEADER_FONT):
    for col, value in enumerate(values, 1):
        cell = sheet.cell(row, col, value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

# Sheet 1: Overview
ws1 = wb.create_sheet('Overview')
set_column_widths(ws1, [20, 15, 15, 15, 15])

ws1['A1'] = '🚀 Supabase 최적화 로드맵'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws1.merge_cells('A1:E1')

data = [
    ['Phase', '규모', 'API 응답', 'DB 응답', '기간'],
    ['Phase 1', '3K → 10K', '<100ms', '<50ms', 'Week 2-3'],
    ['Phase 2', '10K → 100K', '<150ms', '<75ms', 'Month 3-6'],
    ['Phase 3', '100K → 1M', '<200ms', '<100ms', 'Month 6-12'],
]
add_header(ws1, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(i, j, val)
        cell.border = BORDER
        if 'Phase 1' in str(val):
            cell.fill = PHASE1_FILL
        elif 'Phase 2' in str(val):
            cell.fill = PHASE2_FILL
        elif 'Phase 3' in str(val):
            cell.fill = PHASE3_FILL

ws1['A8'] = '비용 예상'
ws1['A8'].font = BOLD_FONT
ws1['A8'].fill = SUBHEADER_FILL

data = [
    ['규모', 'Supabase', 'Redis', 'Replica', '합계/월'],
    ['3K', 'Free', '-', '-', '$0'],
    ['10K', '$25', '$20', '-', '$45'],
    ['100K', '$125', '$50', '$125', '$400'],
    ['1M', '$750', '$200', '$750', '$2,200'],
]
add_header(ws1, 9, data[0])
for i, row in enumerate(data[1:], 10):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(i, j, val)
        cell.border = BORDER

# Sheet 2: Phase 1 Checklist
ws2 = wb.create_sheet('Phase 1 (Week 2-3)')
set_column_widths(ws2, [10, 40, 15, 15, 10])

ws2['A1'] = 'Phase 1: 기본 최적화 (3K → 10K)'
ws2['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws2.merge_cells('A1:E1')

ws2['A3'] = '1. 인덱스 생성 (30개)'
ws2['A3'].font = BOLD_FONT
ws2['A3'].fill = SUBHEADER_FILL
ws2.merge_cells('A3:E3')

data = [
    ['No', '작업', '테이블', '소요시간', '완료'],
    ['1', 'type, status, parent_id, phone 인덱스', 'profiles', '2분', '☐'],
    ['2', 'student_id, status, due_date 인덱스', 'payments', '2분', '☐'],
    ['3', 'student_id, schedule_id, date 인덱스', 'bookings', '2분', '☐'],
    ['4', 'student_id, status, due_date 인덱스', 'invoices', '2분', '☐'],
    ['5', 'invoice_id, paid_at, card_company 인덱스', 'payment_transactions', '2분', '☐'],
]
add_header(ws2, 4, data[0])
for i, row in enumerate(data[1:], 5):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i, j, val)
        cell.border = BORDER

ws2['A11'] = '2. Materialized View (3개)'
ws2['A11'].font = BOLD_FONT
ws2['A11'].fill = SUBHEADER_FILL
ws2.merge_cells('A11:E11')

data = [
    ['No', '작업', 'View명', '소요시간', '완료'],
    ['6', '학생별 미수금 현황', 'mv_student_unpaid_summary', '1분', '☐'],
    ['7', '일별 매출 집계', 'mv_daily_sales', '1분', '☐'],
    ['8', '월별 청구서 현황', 'mv_monthly_invoice_summary', '1분', '☐'],
]
add_header(ws2, 12, data[0])
for i, row in enumerate(data[1:], 13):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i, j, val)
        cell.border = BORDER

ws2['A17'] = '3. 자동화 (pg_cron)'
ws2['A17'].font = BOLD_FONT
ws2['A17'].fill = SUBHEADER_FILL
ws2.merge_cells('A17:E17')

data = [
    ['No', '작업', '실행 주기', '소요시간', '완료'],
    ['9', 'mv_daily_sales 갱신', '매일 03:00', '1분', '☐'],
    ['10', 'mv_monthly_invoice_summary 갱신', '매월 1일 03:00', '1분', '☐'],
    ['11', 'mv_student_unpaid_summary 갱신', '매시간', '1분', '☐'],
    ['12', '만료된 알림 삭제', '매일 02:00', '1분', '☐'],
]
add_header(ws2, 18, data[0])
for i, row in enumerate(data[1:], 19):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i, j, val)
        cell.border = BORDER

ws2['A24'] = '4. RLS 정책'
ws2['A24'].font = BOLD_FONT
ws2['A24'].fill = SUBHEADER_FILL
ws2.merge_cells('A24:E24')

data = [
    ['No', '작업', '테이블', '소요시간', '완료'],
    ['13', 'Service Role 전체 접근', 'All tables', '2분', '☐'],
    ['14', '사용자 본인 데이터만 조회', 'profiles, payments, bookings', '3분', '☐'],
]
add_header(ws2, 25, data[0])
for i, row in enumerate(data[1:], 26):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i, j, val)
        cell.border = BORDER

ws2['A29'] = '총 소요 시간: 약 15분'
ws2['A29'].font = BOLD_FONT

# Sheet 3: Phase 2 Checklist
ws3 = wb.create_sheet('Phase 2 (Month 3-6)')
set_column_widths(ws3, [10, 45, 15, 10])

ws3['A1'] = 'Phase 2: 중급 최적화 (10K → 100K)'
ws3['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws3.merge_cells('A1:D1')

data = [
    ['No', '작업', '소요시간', '완료'],
    ['1', 'PgBouncer 연결 풀링 설정', '1시간', '☐'],
    ['2', 'Redis 캐싱 구현 (5-10분 TTL)', '4시간', '☐'],
    ['3', 'payment_transactions 월별 파티셔닝', '2시간', '☐'],
    ['4', 'FastAPI 병렬 쿼리 적용', '3시간', '☐'],
    ['5', 'Supabase Pooler URL 전환', '30분', '☐'],
]
add_header(ws3, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(i, j, val)
        cell.border = BORDER

ws3['A10'] = '총 소요 시간: 약 10.5시간 (개발자 1-2명)'
ws3['A10'].font = BOLD_FONT

# Sheet 4: Phase 3 Checklist
ws4 = wb.create_sheet('Phase 3 (Month 6-12)')
set_column_widths(ws4, [10, 45, 15, 10])

ws4['A1'] = 'Phase 3: 고급 최적화 (100K → 1M)'
ws4['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws4.merge_cells('A1:D1')

data = [
    ['No', '작업', '소요시간', '완료'],
    ['1', 'Read Replica 설정', '2시간', '☐'],
    ['2', 'Full-Text Search 구현 (profiles)', '4시간', '☐'],
    ['3', '데이터베이스 샤딩 준비 (10 shards)', '1주', '☐'],
    ['4', 'CDC → ClickHouse 연동', '1주', '☐'],
    ['5', 'Load Balancer 설정', '2일', '☐'],
]
add_header(ws4, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(i, j, val)
        cell.border = BORDER

ws4['A10'] = '총 소요 시간: 약 3주 (개발팀 3-5명)'
ws4['A10'].font = BOLD_FONT

# Sheet 5: Monitoring
ws5 = wb.create_sheet('Monitoring')
set_column_widths(ws5, [25, 50, 15])

ws5['A1'] = '모니터링 체크리스트'
ws5['A1'].font = Font(bold=True, size=12, color='1F4E78')
ws5.merge_cells('A1:C1')

data = [
    ['메트릭', '확인 방법', '목표'],
    ['API 응답 시간', 'FastAPI 미들웨어 로깅', '<100ms'],
    ['DB 쿼리 시간', 'pg_stat_statements', '<50ms'],
    ['느린 쿼리', 'mean_time > 100ms 감지', '0건'],
    ['테이블 크기', 'pg_total_relation_size', '<10GB'],
    ['인덱스 사용률', 'idx_scan > 100', '>90%'],
    ['캐시 히트율', 'Redis INFO stats', '>80%'],
    ['동시 접속', 'pg_stat_activity', '<100'],
    ['CPU 사용률', 'Supabase Dashboard', '<70%'],
    ['메모리 사용률', 'Supabase Dashboard', '<80%'],
]
add_header(ws5, 3, data[0])
for i, row in enumerate(data[1:], 4):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

ws5['A14'] = '알람 설정'
ws5['A14'].font = BOLD_FONT
ws5['A14'].fill = SUBHEADER_FILL

data = [
    ['조건', '알림 채널', '담당'],
    ['API 응답 > 500ms', '몰트봇', '개발팀'],
    ['DB 연결 > 80개', '이메일', '인프라'],
    ['테이블 크기 > 10GB', '이메일', '인프라'],
    ['에러율 > 1%', '몰트봇', '개발팀'],
]
add_header(ws5, 15, data[0])
for i, row in enumerate(data[1:], 16):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(i, j, val)
        cell.border = BORDER

wb.save('/sessions/modest-bold-einstein/mnt/autus/Supabase_최적화_체크리스트.xlsx')
print("✅ Excel 파일 생성 완료")
