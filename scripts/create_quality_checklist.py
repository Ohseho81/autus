#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD_FONT = Font(bold=True)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def set_widths(sheet, widths):
    from openpyxl.utils import get_column_letter
    for idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

# Sheet 1: 품질 지표
ws1 = wb.create_sheet('품질 지표')
set_widths(ws1, [25, 20, 15, 15])

ws1['A1'] = '🎯 AUTUS 품질 지표 (7개)'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws1.merge_cells('A1:D1')

data = [
    ['지표', '목표', '현재', '상태'],
    ['정확성 - 출석', '100%', '100%', '✅'],
    ['정확성 - 결제', '100%', '100%', '✅'],
    ['정확성 - 미수금', '100%', '100%', '✅'],
    ['신뢰성 - Uptime', '>99.9%', '-', '⏳'],
    ['사용성 - 만족도', '>85%', '-', '⏳'],
    ['성능 - API P95', '<200ms', '-', '⏳'],
    ['보안 - 데이터 유출', '0건', '0건', '✅'],
]

for i, row in enumerate(data):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(i+3, j, val)
        cell.border = BORDER
        if i == 0:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif '✅' in str(val):
            cell.fill = SUCCESS_FILL
        elif '⏳' in str(val):
            cell.fill = WARNING_FILL

# Sheet 2: Week 2 체크리스트
ws2 = wb.create_sheet('Week 2 (기초)')
set_widths(ws2, [40, 15, 10])

ws2['A1'] = 'Week 2 품질 체크리스트'
ws2['A1'].font = Font(bold=True, size=12)
ws2.merge_cells('A1:C1')

data = [
    ['항목', '소요시간', '완료'],
    ['단위 테스트 커버리지 >80%', '1일', '☐'],
    ['출석/결제/미수금 통합 테스트', '0.5일', '☐'],
    ['API 응답 <200ms 검증', '0.5일', '☐'],
    ['RLS 정책 100% 적용', '0.5일', '☐'],
    ['에러 핸들링 모든 API', '1일', '☐'],
    ['로깅 시스템 구축', '0.5일', '☐'],
    ['몰트봇 알림 연동', '0.5일', '☐'],
    ['Health Check 엔드포인트', '0.5일', '☐'],
]

for i, row in enumerate(data):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(i+3, j, val)
        cell.border = BORDER
        if i == 0:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

ws2['A13'] = '총 소요: 5일'
ws2['A13'].font = BOLD_FONT

wb.save('/sessions/modest-bold-einstein/mnt/autus/품질_체크리스트.xlsx')
print("✅ 품질 체크리스트 생성 완료")
